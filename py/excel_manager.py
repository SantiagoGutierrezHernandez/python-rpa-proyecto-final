from openpyxl import load_workbook
from py.invoice_manager import InvoiceItem

class ExcelWrapper:
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(filename, data_only=True)
        self.ws = self.wb.active

    def iter_rows(self, min_row = 2, max_row = None, min_col = 1, max_col = None):
        if max_row is None: max_row = self.ws.max_row
        if max_col is None: max_col = self.ws.max_column
        
        return self.ws.iter_rows(min_row, max_row, min_col, max_col)

    def set_value(self, cell, value):
        cell.value = value
        self.wb.save(self.filename)

    @staticmethod
    def to_dict(row, index_to_key_dict = {0:"nombre"}):
        new_dict = {}
        for index, keyname in index_to_key_dict.items():
            new_dict[keyname] = row[index].value
        return new_dict

    @staticmethod
    def to_item(row):
        item = ExcelWrapper.to_dict(row, {
            0:"numero",
            1:"comprador",
            2:"id",
            3:"fecha",
            4:"monto",
            5:"cantidad"
        })
        return InvoiceItem(item.get("id"), item.get("comprador"), item.get("monto"), item.get("cantidad"), item.get("fecha"), item.get("numero"))